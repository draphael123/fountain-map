"""One-off export from Provider Compliance Dashboard xlsx -> src/data/*.ts"""
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("pip install openpyxl")

XLSX = Path(r"c:\Users\danie\Downloads\Provider _ Compliance Dashboard (31).xlsx")
OUT = Path(__file__).resolve().parent.parent / "src" / "data"

NAME_TO_ID = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR", "california": "CA",
    "colorado": "CO", "connecticut": "CT", "delaware": "DE", "florida": "FL", "georgia": "GA",
    "hawaii": "HI", "idaho": "ID", "illinois": "IL", "indiana": "IN", "iowa": "IA", "kansas": "KS",
    "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD", "massachusetts": "MA",
    "michigan": "MI", "minnesota": "MN", "mississippi": "MS", "missouri": "MO", "montana": "MT",
    "nebraska": "NE", "nevada": "NV", "new hampshire": "NH", "new jersey": "NJ", "new mexico": "NM",
    "new york": "NY", "north carolina": "NC", "north dakota": "ND", "ohio": "OH", "oklahoma": "OK",
    "oregon": "OR", "pennsylvania": "PA", "rhode island": "RI", "south carolina": "SC",
    "south dakota": "SD", "tennessee": "TN", "texas": "TX", "utah": "UT", "vermont": "VT",
    "virginia": "VA", "washington": "WA", "west virginia": "WV", "wisconsin": "WI", "wyoming": "WY",
    "district of columbia": "DC", "washington d.c.": "DC", "washington dc": "DC",
}


def resolve_state_id(label: str) -> str | None:
    if not label:
        return None
    s = label.strip()
    m = re.search(r"\(([A-Z]{2})\)\s*$", s)
    if m:
        return m.group(1)
    key = re.sub(r"\s+", " ", s.lower())
    key = key.replace(".", "").replace("(", "").replace(")", "")
    if key in NAME_TO_ID:
        return NAME_TO_ID[key]
    # strip trailing (XX) already handled; try first part before comma
    first = key.split(",")[0].strip()
    return NAME_TO_ID.get(first)


def esc(s: str) -> str:
    if s is None:
        return ""
    return (
        str(s)
        .replace("\\", "\\\\")
        .replace("'", "\\'")
        .replace("\n", " ")
        .replace("\r", "")
    )


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    # State Rules
    ws = wb["State Rules- NEW"]
    rules = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        st = str(row[0]).strip()
        if st.lower() == "state":
            continue
        sid = resolve_state_id(st)
        if not sid:
            continue
        rules.append(
            {
                "stateId": sid,
                "stateLabel": st,
                "compact": esc(row[1]) if row[1] else "",
                "operational": esc(row[2]) if row[2] else "",
                "requiredSteps": esc(row[3]) if row[3] else "",
                "deaCsrs": esc(row[4]) if row[4] else "",
                "cpaRequired": esc(row[5]) if row[5] else "",
                "notes": esc(row[6]) if row[6] else "",
            }
        )

    rules.sort(key=lambda x: x["stateId"])

    lines = [
        "/** State licensing rules — exported from Provider Compliance Dashboard (State Rules- NEW) */",
        "export interface StateRuleRow {",
        "  stateId: string;",
        "  stateLabel: string;",
        "  compact: string;",
        "  operational: string;",
        "  requiredSteps: string;",
        "  deaCsrs: string;",
        "  cpaRequired: string;",
        "  notes: string;",
        "}",
        "",
        "export const STATE_RULES_ROWS: StateRuleRow[] = [",
    ]
    for r in rules:
        lines.append(
            "  { "
            + ", ".join(
                f"{k}: '{v}'" if k != "stateId" else f'{k}: "{v}"'
                for k, v in r.items()
            )
            + " },"
        )
    lines.append("];")
    lines.append("")
    (OUT / "stateRulesData.ts").write_text("\n".join(lines), encoding="utf-8")
    print("Wrote stateRulesData.ts", len(rules), "rows")

    # CS Refill
    ws2 = wb["State CS Refill Guidelines"]
    refills = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        st = str(row[0]).strip()
        if st.lower() == "state":
            continue
        sid = resolve_state_id(st)
        if not sid:
            continue
        refills.append(
            {
                "stateId": sid,
                "stateLabel": st,
                "restriction": esc(row[1]) if row[1] else "",
                "extraNotes": esc(row[3]) if len(row) > 3 and row[3] else "",
            }
        )
    refills.sort(key=lambda x: x["stateId"])

    lines2 = [
        "/** CS prescription/refill notes by state — exported from Provider Compliance Dashboard */",
        "export interface StateCSRefillRow {",
        "  stateId: string;",
        "  stateLabel: string;",
        "  restriction: string;",
        "  extraNotes: string;",
        "}",
        "",
        "export const STATE_CS_REFILL_ROWS: StateCSRefillRow[] = [",
    ]
    for r in refills:
        lines2.append(
            "  { "
            + ", ".join(
                f"{k}: '{v}'" if k != "stateId" else f'{k}: "{v}"'
                for k, v in r.items()
            )
            + " },"
        )
    lines2.append("];")
    (OUT / "stateCSRefillData.ts").write_text("\n".join(lines2), encoding="utf-8")
    print("Wrote stateCSRefillData.ts", len(refills), "rows")


if __name__ == "__main__":
    main()
